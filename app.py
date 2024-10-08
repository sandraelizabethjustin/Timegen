
from flask import Flask, render_template, request,send_file
from flask import Flask, send_from_directory
import pandas as pd
import random
import xlsxwriter
import openpyxl
classes=[]
app = Flask(__name__)
TOTAL_HRS=7
DAYS=5
MAX_SIZE=120
GAP=17

def populate_teacher(s):
    length=len(s)
    list_1=[]
    class_ind={}
    for k in range(length):
        if str(s.iat[k,1]).lower()!='nan':
            class_ind[s.iat[k,0]]=s.iat[k,1].split(',')
            #class_ind.append(list((s.iat[k,0],s.iat[k,1])))
        else:
            classes.append(str(s.iat[k,0]))
    
    return class_ind

def populate(s):
    length=len(s)
    list_1=[]
    class_ind=[]
    for k in range(length):
        if str(s.iat[k,1]).lower()!='nan':
            class_ind.append(list((s.iat[k,0],s.iat[k,1])))
        else:
            if  class_ind:
                list_1.append(class_ind)
            classes.append(str(s.iat[k,0]))
            class_ind=[]
    list_1.append(class_ind)
    return list_1

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/page2.html')
def page2():
    return render_template('page2.html')

@app.route('/downloads/<path:filename>')
def download_file(filename):
    
    return send_file('static/' + filename, as_attachment=True)


@app.route('/view', methods=['POST'])
def view():
    files = [request.files[f'file{i}'] for i in range(1, 4)]
   # file1.save(file1.filename)
    s2=pd.read_excel(files[0],skiprows=2)
    s1=pd.read_excel(files[1])
    s3=pd.read_excel(files[2])
    
    path = files[0] 
    wb_obj = openpyxl.load_workbook(path)   
    sheet_obj = wb_obj.active 
    cell_obj1 = str(sheet_obj.cell(row=1, column=1).value)
    cell_obj2 =str( sheet_obj.cell(row=2, column=1).value)
    
    #s1=pd.read_excel("Course_teacher_map.xlsx")
    wb = xlsxwriter.Workbook('static/final.xlsx')
    ws = wb.add_worksheet("TimeTable")
    ws2=wb.add_worksheet("TeacherSlot")
    f2= wb.add_format({'bold':True,'bg_color':'#b2b2b2'})
    f3=wb.add_format({'bg_color':'#808080'})
    f4=wb.add_format({'bold':True,'bg_color':'#808080'})
    f5=wb.add_format({'bg_color':'#b2b2b2'})
    f6 = wb.add_format({'bold':True,'bg_color':'#999999'})
    f7=wb.add_format({'bold':True})
    working_days=["Monday","Tuesday","Wednesday","Thursday","Friday"]



    teachers=s1['faculty'].dropna().unique().tolist()
    tt=[]
    for teacher in teachers:
        a=teacher.split(',')
        if len(a)==1:
            tt.append(a[0])
        else:
            for i in a:
                tt.append(i)
    teachers=list(set(tt))
    print(teachers)

    teacher_course=populate_teacher(s1)
    print(teacher_course)
    teacher_len=len(teachers)

    #s2=pd.read_excel("ti.xlsx")
    t_len=len(s2)
    print(t_len)
    #s3=pd.read_excel("course_hour_map.xlsx")
    course_hour=populate(s3)
    print(course_hour)
    timeslot=[[0]*MAX_SIZE for i in range(t_len)]
    teacherslot=[[0]*MAX_SIZE for i in range(teacher_len)]

    for i in range(t_len):
        index=0
        for k in range(DAYS):
            for j in range(TOTAL_HRS):
                timeslot[i][index]=str(s2.iat[i,(k*TOTAL_HRS)+j])
                if timeslot[i][index] in teacher_course:
                    fac=teacher_course[timeslot[i][index]]
                    t_index=[]
                    for teacher in fac:
                        t_index.append(teachers.index(teacher))
                    for tindex in t_index:
                         teacherslot[tindex][index]=timeslot[i][index]
                index+=1
            index+=GAP


    for k in range(t_len):
        #c_t=teacher_course[k]
        c_h=course_hour[k]
        for i in range(len(c_h)):
            course=c_h[i][0]
            hour=c_h[i][1]
            #faculty=[inner_list[1] for inner_list in c_t if inner_list[0]==course]
            #fac=faculty[0]
            #t_index=teachers.index(fac)
            fac=teacher_course[course]
            t_index=[]
            for teacher in fac:
                t_index.append(teachers.index(teacher))
            alloc_hr=0
            rem_hr=hour
            slots=[]
            while int(rem_hr)>0:
                for j in range(MAX_SIZE):
                    if str(timeslot[k][j]).lower()=="nan":
                        begin=j
                        break
                for j in range((MAX_SIZE-1),-1,-1):
                    if str(timeslot[k][j]).lower()=="nan":
                        end=j
                        break
                if rem_hr!=1:
                    interval=(end-begin+1)/(rem_hr-1)
                pos=begin
                for j in range(int(rem_hr)):
                    slots.append(pos)
                    pos=pos+interval
                for slot in slots:
                    flag=0
                    flag2=0
                    if str(timeslot[k][int(slot)]).lower()=="nan" :
                        for tindex in t_index:
                            if teacherslot[tindex][int(slot)]!=0 or teacherslot[tindex][(int(slot)-1)%MAX_SIZE]!=0 or teacherslot[tindex][(int(slot)+1)%MAX_SIZE]!=0:
                                flag=1
                            
                        if flag==0:
                            timeslot[k][int(slot)]=course
                            for tindex in t_index:
                                teacherslot[tindex][int(slot)]=course
                            flag2=1
                        
                    elif flag2==0:
                        left=int(slot)-1
                        right=int(slot)+1
                        while left>0 or right<MAX_SIZE:
                            if((left>0 and str(timeslot[k][left]).lower()=="nan")) :
                                f=0
                                for tindex in t_index:
                                    if teacherslot[tindex][left]!=0 or teacherslot[tindex][(left-1)%MAX_SIZE]!=0 or teacherslot[tindex][(left+1)%MAX_SIZE]:
                                        f=1
                                if f==0:        
                                    timeslot[k][left]=course
                                    for tindex in t_index:
                                        teacherslot[tindex][left]=course
                                    break
                            if(right<MAX_SIZE and str(timeslot[k][right]).lower()=="nan") :
                                f=0
                                for tindex in t_index:
                                    if teacherslot[tindex][right]!=0 or teacherslot[tindex][(right-1)%MAX_SIZE]!=0 or teacherslot[tindex][(right+1)%MAX_SIZE]:
                                        f=1
                                if f==0:  
                                    timeslot[k][right]=course
                                    for tindex in t_index:
                                        teacherslot[tindex][right]=course
                                    break
                            left=left-1
                            right=right+1
                        if left<0 and right>=MAX_SIZE:
                            print("ERROR: ALLOCATION COULD NOT BE DONE")
                            break
                            
                    rem_hr-=1
                else:
                    continue
                break       
            
    k=0
    timetable=[]
    counter=3
    
    merge_format = wb.add_format(
    {
        "bold": 1,
        "align": "center",
        "valign": "vcenter"
    }
    )
    ws.merge_range("A1:U1", cell_obj1, merge_format)
    ws.merge_range("A2:U2", cell_obj2, merge_format)
    
    while k<t_len:
        index=0
        temp=[[0]*TOTAL_HRS for i in range(DAYS)]
        timetable.append(['','','',classes[k],'','',''])
        for i in range(DAYS):
            for j in range(TOTAL_HRS):
                if str(timeslot[k][index]).lower()=="nan":
                    timeslot[k][index] ="REMEDIAL"
                temp[i][j]=timeslot[k][index]           
                index+=1
            timetable.append(temp[i])
            index+=GAP
        ws.write(counter,4,classes[k])
        if k%2==0:
            ws.write_row(counter+1,0,['','1st','2nd','3rd','Lunch','4th','5th','6th'],f6)
        else:
            ws.write_row(counter+1,0,['','1st','2nd','3rd','Lunch','4th','5th','6th'],f4)
        for row in range(len(temp)):
        # ws2.write_row(counter+row,1,['1st','2nd','3rd','Lunch','4th','5th','6th'],f2)
            if row%2==0:
                ws.write(counter + row+2, 0, working_days[row],f6)
            else:
                ws.write(counter + row+2, 0, working_days[row],f4)
            
            for col, value in enumerate(temp[row]):
                if row%2==0:
                    ws.write(counter + row+2, col + 1, value,f5)
                else:
                    ws.write(counter + row+2, col + 1, value,f3)
        counter+=8
        k=k+1
        
   # print(timetable)

    k=0
    teachslot=[]
    counter=3
    

    ws2.merge_range("A1:U1", cell_obj1, merge_format)
    ws2.merge_range("A2:U2", cell_obj2, merge_format)
    
    while k<teacher_len:
        index=0
        temp=[[0]*TOTAL_HRS for i in range(DAYS)]
        teachslot.append([teachers[k],'','','','','',''])
        for i in range(DAYS):
            for j in range(TOTAL_HRS):
                if teacherslot[k][index]==0:
                    teacherslot[k][index]="-"
                temp[i][j]=teacherslot[k][index]
                index+=1
            teachslot.append(temp[i])    
            index+=GAP
        ws2.write(counter,4,teachers[k],f7)
        if k%2==0:
            ws2.write_row(counter+1,0,['','1st','2nd','3rd','Lunch','4th','5th','6th'],f6)
        else:
            ws2.write_row(counter+1,0,['','1st','2nd','3rd','Lunch','4th','5th','6th'],f4)
        for row in range(len(temp)):
        # ws2.write_row(counter+row,1,['1st','2nd','3rd','Lunch','4th','5th','6th'],f2)
            if row%2==0:
                ws2.write(counter + row+2, 0, working_days[row],f6)
            else:
                ws2.write(counter + row+2, 0, working_days[row],f4)
            for col, value in enumerate(temp[row]):
                if row%2==0:
                    ws2.write(counter + row+2, col + 1, value,f5)
                else:
                    ws2.write(counter + row+2, col + 1, value,f3)
        counter+=8
        k=k+1
    #print(teachslot)
    wb.close()


    # tf=pd.DataFrame(timetable,index=['','monday','tuesday','wednesday','thursday','friday']*t_len,columns=['1st','2nd','3rd','Lunch','4th','5th','6th'])
    # #tf.to_excel("static/tf7.xlsx")
    # #return send_file("static/tf7.xlsx", as_attachment=True)
    # tc=pd.DataFrame(teachslot,index=['FACULTY','MONDAY','TUESDAY','WEDNESDAY','THURSDAY','FRIDAY']*teacher_len,columns=['1st','2nd','3rd','Lunch','4th','5th','6th'])
    # #tc.to_excel("static/tc3.xlsx")
    # #return send_file("static/tc3.xlsx", as_attachment=True)
    # #return send_file("static/tf7.xlsx", as_attachment=True), send_file("static/tc3.xlsx", as_attachment=True)
    # with pd.ExcelWriter("static/final.xlsx") as writer:
    #     tf.to_excel(writer,sheet_name="Timetable")
    #     tc.to_excel(writer,sheet_name="Teacher Slots")#ts
    return send_file("static/final.xlsx", as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)